#!/usr/bin/env python3
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSX = ROOT / "Orçamento mensal.xlsx"
TEMPLATE_XLSX = ROOT / "Orcamento_Robusto_Template.xlsx"
FILLED_XLSX = ROOT / "Orcamento_Robusto_Preenchido.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


@dataclass
class ExtractedData:
    expense_categories: list[tuple[str, float | None]]
    income_categories: list[tuple[str, float | None]]
    subscriptions: list[dict[str, Any]]
    transactions: list[dict[str, Any]]


def normalize_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass
    return None


def extract_current_data(source_path: Path) -> ExtractedData:
    wb = load_workbook(source_path, data_only=True)

    resumo = wb["Resumo"]
    transacoes = wb["Transações"]
    assinaturas = wb["Assinaturas"]

    expense_categories: list[tuple[str, float | None]] = []
    for row in range(28, 44):
        category = resumo.cell(row, 2).value
        planned = resumo.cell(row, 4).value
        if not category:
            continue
        if isinstance(planned, (int, float)):
            expense_categories.append((str(category).strip(), float(planned)))
        else:
            expense_categories.append((str(category).strip(), None))

    income_categories: list[tuple[str, float | None]] = []
    for row in range(28, 44):
        category = resumo.cell(row, 8).value
        planned = resumo.cell(row, 10).value
        if not category:
            continue
        income_categories.append((str(category).strip(), float(planned) if isinstance(planned, (int, float)) else None))

    subscriptions: list[dict[str, Any]] = []
    freq = "monthly"
    sub_idx = 1
    for row in range(1, assinaturas.max_row + 1):
        name = assinaturas.cell(row, 1).value
        amount = assinaturas.cell(row, 2).value
        if not name:
            continue
        normalized_name = str(name).strip()
        lowered = normalized_name.lower()
        if "assinaturas anuais" in lowered:
            freq = "yearly"
            continue
        if lowered == "assinaturas":
            freq = "monthly"
            continue
        if not isinstance(amount, (int, float)):
            continue
        subscriptions.append(
            {
                "subscription_id": f"SUB-{sub_idx:03d}",
                "name": normalized_name,
                "amount": float(amount),
                "frequency": freq,
                "start_date": date.today(),
                "day_of_month": 1,
                "category": "Assinaturas",
                "active": "yes",
                "notes": "Importado da planilha original",
            }
        )
        sub_idx += 1

    transactions: list[dict[str, Any]] = []
    tx_idx = 1
    for row in range(5, transacoes.max_row + 1):
        exp_date = normalize_date(transacoes.cell(row, 2).value)
        exp_amount = transacoes.cell(row, 3).value
        exp_desc = transacoes.cell(row, 4).value
        exp_cat = transacoes.cell(row, 5).value
        if exp_date and isinstance(exp_amount, (int, float)) and exp_desc and exp_cat:
            transactions.append(
                {
                    "transaction_id": f"TX-{tx_idx:04d}",
                    "date": exp_date,
                    "nature": "expense",
                    "entry_type": "one_off",
                    "category": str(exp_cat).strip(),
                    "description": str(exp_desc).strip(),
                    "amount": float(exp_amount),
                    "reference_id": "",
                    "account": "",
                    "notes": "Importado da aba Transações",
                }
            )
            tx_idx += 1

        inc_date = normalize_date(transacoes.cell(row, 7).value)
        inc_amount = transacoes.cell(row, 8).value
        inc_desc = transacoes.cell(row, 9).value
        inc_cat = transacoes.cell(row, 10).value
        if inc_date and isinstance(inc_amount, (int, float)) and inc_desc and inc_cat:
            transactions.append(
                {
                    "transaction_id": f"TX-{tx_idx:04d}",
                    "date": inc_date,
                    "nature": "income",
                    "entry_type": "one_off",
                    "category": str(inc_cat).strip(),
                    "description": str(inc_desc).strip(),
                    "amount": float(inc_amount),
                    "reference_id": "",
                    "account": "",
                    "notes": "Importado da aba Transações",
                }
            )
            tx_idx += 1

    return ExtractedData(
        expense_categories=expense_categories,
        income_categories=income_categories,
        subscriptions=subscriptions,
        transactions=transactions,
    )


def style_headers(ws, header_row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(header_row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def add_table(ws, name: str, start_row: int, end_row: int, end_col: int) -> None:
    ref = f"A{start_row}:{chr(64 + end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def ensure_table_row_count(data_len: int) -> int:
    return max(2, data_len + 1)


def build_workbook(output_path: Path, extracted: ExtractedData | None) -> None:
    wb = Workbook()

    ws_intro = wb.active
    ws_intro.title = "Instrucoes"
    ws_intro.append(["Planilha de Controle Financeiro (Estrutura Robusta)"])
    ws_intro.append([
        "Fluxo recomendado: 1) configure Categorias/Assinaturas/Orcamento, 2) lance em Transacoes, 3) veja Dashboard."
    ])
    ws_intro.append([
        "Use IDs estáveis (ex.: SUB-001, PAR-001) para rastrear lançamentos automáticos e parcelas."
    ])
    ws_intro.append([
        "Não apague colunas. Pode adicionar novas linhas nas tabelas (elas se expandem automaticamente)."
    ])
    ws_intro.column_dimensions["A"].width = 130
    ws_intro["A1"].font = Font(size=14, bold=True)

    ws_lists = wb.create_sheet("Listas")
    ws_lists.append(["nature", "entry_type", "frequency", "yes_no"])
    max_list_len = 4
    values = [
        ["expense", "one_off", "monthly", "yes"],
        ["income", "subscription", "yearly", "no"],
        ["", "installment", "", ""],
        ["", "", "", ""],
    ]
    for row in values:
        ws_lists.append(row)
    ws_lists.sheet_state = "hidden"

    ws_cat = wb.create_sheet("Categorias")
    cat_headers = ["category_id", "nature", "category_name", "active", "notes"]
    ws_cat.append(cat_headers)

    category_rows: list[list[Any]] = []
    if extracted:
        seen = set()
        idx = 1
        for cat, _ in extracted.expense_categories:
            key = ("expense", cat)
            if key in seen:
                continue
            seen.add(key)
            category_rows.append([f"CAT-{idx:03d}", "expense", cat, "yes", "Importado"])
            idx += 1
        for cat, _ in extracted.income_categories:
            key = ("income", cat)
            if key in seen:
                continue
            seen.add(key)
            category_rows.append([f"CAT-{idx:03d}", "income", cat, "yes", "Importado"])
            idx += 1

    if not category_rows:
        category_rows = [["CAT-001", "expense", "", "yes", ""]]

    for row in category_rows:
        ws_cat.append(row)

    style_headers(ws_cat, 1, len(cat_headers))
    ws_cat.column_dimensions["A"].width = 14
    ws_cat.column_dimensions["B"].width = 12
    ws_cat.column_dimensions["C"].width = 28
    ws_cat.column_dimensions["D"].width = 10
    ws_cat.column_dimensions["E"].width = 36
    add_table(ws_cat, "tbl_categories", 1, ensure_table_row_count(len(category_rows)), len(cat_headers))

    ws_budget = wb.create_sheet("Orcamento")
    budget_headers = ["month", "nature", "category_name", "planned_amount", "notes"]
    ws_budget.append(budget_headers)

    budget_rows: list[list[Any]] = []
    if extracted:
        # Use current month as default planning cycle for imported baseline values.
        month_value = date.today().replace(day=1)
        for cat, planned in extracted.expense_categories:
            if planned is None:
                continue
            budget_rows.append([month_value, "expense", cat, planned, "Planejado da planilha base"])
        for cat, planned in extracted.income_categories:
            if planned is None:
                continue
            budget_rows.append([month_value, "income", cat, planned, "Planejado da planilha base"])

    if not budget_rows:
        budget_rows = [[date.today().replace(day=1), "expense", "", 0.0, ""]]

    for row in budget_rows:
        ws_budget.append(row)
    style_headers(ws_budget, 1, len(budget_headers))
    ws_budget.column_dimensions["A"].width = 14
    ws_budget.column_dimensions["B"].width = 12
    ws_budget.column_dimensions["C"].width = 28
    ws_budget.column_dimensions["D"].width = 16
    ws_budget.column_dimensions["E"].width = 38
    add_table(ws_budget, "tbl_budget", 1, ensure_table_row_count(len(budget_rows)), len(budget_headers))

    ws_sub = wb.create_sheet("Assinaturas")
    sub_headers = [
        "subscription_id",
        "name",
        "amount",
        "frequency",
        "start_date",
        "day_of_month",
        "category_name",
        "active",
        "notes",
    ]
    ws_sub.append(sub_headers)

    sub_rows = []
    if extracted:
        for sub in extracted.subscriptions:
            sub_rows.append(
                [
                    sub["subscription_id"],
                    sub["name"],
                    sub["amount"],
                    sub["frequency"],
                    sub["start_date"],
                    sub["day_of_month"],
                    sub["category"],
                    sub["active"],
                    sub["notes"],
                ]
            )

    if not sub_rows:
        sub_rows = [["SUB-001", "", 0.0, "monthly", date.today(), 1, "Assinaturas", "yes", ""]]

    for row in sub_rows:
        ws_sub.append(row)

    style_headers(ws_sub, 1, len(sub_headers))
    ws_sub.column_dimensions["A"].width = 16
    ws_sub.column_dimensions["B"].width = 24
    ws_sub.column_dimensions["C"].width = 12
    ws_sub.column_dimensions["D"].width = 12
    ws_sub.column_dimensions["E"].width = 12
    ws_sub.column_dimensions["F"].width = 14
    ws_sub.column_dimensions["G"].width = 20
    ws_sub.column_dimensions["H"].width = 10
    ws_sub.column_dimensions["I"].width = 36
    add_table(ws_sub, "tbl_subscriptions", 1, ensure_table_row_count(len(sub_rows)), len(sub_headers))

    ws_inst = wb.create_sheet("Parcelados")
    inst_headers = [
        "installment_id",
        "description",
        "total_installments",
        "total_amount",
        "purchase_date",
        "first_due_date",
        "category_name",
        "store",
        "active",
        "notes",
    ]
    ws_inst.append(inst_headers)
    inst_rows = [["PAR-001", "", 12, 0.0, date.today(), date.today(), "", "", "yes", ""]]
    for row in inst_rows:
        ws_inst.append(row)

    style_headers(ws_inst, 1, len(inst_headers))
    for col, width in {
        "A": 16,
        "B": 28,
        "C": 18,
        "D": 14,
        "E": 12,
        "F": 14,
        "G": 20,
        "H": 20,
        "I": 10,
        "J": 34,
    }.items():
        ws_inst.column_dimensions[col].width = width
    add_table(ws_inst, "tbl_installments", 1, ensure_table_row_count(len(inst_rows)), len(inst_headers))

    ws_tx = wb.create_sheet("Transacoes")
    tx_headers = [
        "transaction_id",
        "date",
        "nature",
        "entry_type",
        "category_name",
        "description",
        "amount",
        "reference_id",
        "account",
        "notes",
    ]
    ws_tx.append(tx_headers)

    tx_rows = []
    if extracted:
        for tx in extracted.transactions:
            tx_rows.append(
                [
                    tx["transaction_id"],
                    tx["date"],
                    tx["nature"],
                    tx["entry_type"],
                    tx["category"],
                    tx["description"],
                    tx["amount"],
                    tx["reference_id"],
                    tx["account"],
                    tx["notes"],
                ]
            )
    if not tx_rows:
        tx_rows = [["TX-0001", date.today(), "expense", "one_off", "", "", 0.0, "", "", ""]]

    for row in tx_rows:
        ws_tx.append(row)

    style_headers(ws_tx, 1, len(tx_headers))
    for col, width in {
        "A": 16,
        "B": 12,
        "C": 12,
        "D": 14,
        "E": 24,
        "F": 36,
        "G": 12,
        "H": 16,
        "I": 16,
        "J": 36,
    }.items():
        ws_tx.column_dimensions[col].width = width
    add_table(ws_tx, "tbl_transactions", 1, ensure_table_row_count(len(tx_rows)), len(tx_headers))

    ws_dash = wb.create_sheet("Dashboard")
    ws_dash["A1"] = "Mês de referência (primeiro dia do mês)"
    ws_dash["B1"] = date.today().replace(day=1)
    ws_dash["A3"] = "Receita do mês"
    ws_dash["A4"] = "Despesa do mês"
    ws_dash["A5"] = "Saldo do mês"
    ws_dash["A6"] = "Taxa de economia"
    ws_dash["A8"] = "Orçamento de Despesas"
    ws_dash["A9"] = "Orçamento de Receitas"
    ws_dash["A10"] = "Diferença despesa real vs orçamento"

    ws_dash["B3"] = '=SUMIFS(tbl_transactions[amount],tbl_transactions[nature],"income",tbl_transactions[date],">="&$B$1,tbl_transactions[date],"<="&EOMONTH($B$1,0))'
    ws_dash["B4"] = '=SUMIFS(tbl_transactions[amount],tbl_transactions[nature],"expense",tbl_transactions[date],">="&$B$1,tbl_transactions[date],"<="&EOMONTH($B$1,0))'
    ws_dash["B5"] = "=B3-B4"
    ws_dash["B6"] = '=IFERROR(B5/B3,0)'
    ws_dash["B8"] = '=SUMIFS(tbl_budget[planned_amount],tbl_budget[nature],"expense",tbl_budget[month],$B$1)'
    ws_dash["B9"] = '=SUMIFS(tbl_budget[planned_amount],tbl_budget[nature],"income",tbl_budget[month],$B$1)'
    ws_dash["B10"] = "=B8-B4"

    ws_dash["B6"].number_format = "0.00%"
    for r in (3, 4, 5, 8, 9, 10):
        ws_dash[f"B{r}"].number_format = '#,##0.00'
    ws_dash.column_dimensions["A"].width = 42
    ws_dash.column_dimensions["B"].width = 22
    ws_dash["A1"].font = Font(bold=True)

    date_fmt = "yyyy-mm-dd"
    money_fmt = '#,##0.00'
    month_fmt = "yyyy-mm"

    for ws, col in [(ws_budget, "A"), (ws_sub, "E"), (ws_inst, "E"), (ws_inst, "F"), (ws_tx, "B")]:
        for r in range(2, ws.max_row + 1):
            ws[f"{col}{r}"].number_format = date_fmt
    for r in range(2, ws_budget.max_row + 1):
        ws_budget[f"A{r}"].number_format = month_fmt
    for ws, col in [(ws_budget, "D"), (ws_sub, "C"), (ws_inst, "D"), (ws_tx, "G")]:
        for r in range(2, ws.max_row + 1):
            ws[f"{col}{r}"].number_format = money_fmt

    dv_cat_nature = DataValidation(type="list", formula1="=Listas!$A$2:$A$3", allow_blank=True)
    ws_cat.add_data_validation(dv_cat_nature)
    dv_cat_nature.add("B2:B2000")

    dv_budget_nature = DataValidation(type="list", formula1="=Listas!$A$2:$A$3", allow_blank=True)
    ws_budget.add_data_validation(dv_budget_nature)
    dv_budget_nature.add("B2:B2000")

    dv_tx_nature = DataValidation(type="list", formula1="=Listas!$A$2:$A$3", allow_blank=True)
    ws_tx.add_data_validation(dv_tx_nature)
    dv_tx_nature.add("C2:C2000")

    dv_entry_type = DataValidation(type="list", formula1="=Listas!$B$2:$B$4", allow_blank=True)
    ws_tx.add_data_validation(dv_entry_type)
    dv_entry_type.add("D2:D2000")

    dv_freq = DataValidation(type="list", formula1="=Listas!$C$2:$C$3", allow_blank=True)
    ws_sub.add_data_validation(dv_freq)
    dv_freq.add("D2:D2000")

    dv_cat_yes_no = DataValidation(type="list", formula1="=Listas!$D$2:$D$3", allow_blank=True)
    ws_cat.add_data_validation(dv_cat_yes_no)
    dv_cat_yes_no.add("D2:D2000")

    dv_sub_yes_no = DataValidation(type="list", formula1="=Listas!$D$2:$D$3", allow_blank=True)
    ws_sub.add_data_validation(dv_sub_yes_no)
    dv_sub_yes_no.add("H2:H2000")

    dv_inst_yes_no = DataValidation(type="list", formula1="=Listas!$D$2:$D$3", allow_blank=True)
    ws_inst.add_data_validation(dv_inst_yes_no)
    dv_inst_yes_no.add("I2:I2000")

    wb.save(output_path)


def main() -> int:
    extracted = extract_current_data(SOURCE_XLSX)
    build_workbook(TEMPLATE_XLSX, None)
    build_workbook(FILLED_XLSX, extracted)
    print(f"Generated: {TEMPLATE_XLSX.name}")
    print(f"Generated: {FILLED_XLSX.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
